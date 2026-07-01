# -*- coding: utf-8 -*-
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from config import STOCKS, PROVIDERS
from fetch_market import fetch_with_akshare
from decision_engine import recommend

BASE = Path(__file__).resolve().parent
DATA_DIR = BASE / "data"
REPORT_DIR = BASE / "reports"
EXCEL_PATH = DATA_DIR / "T0_每日记录.xlsx"
DATA_DIR.mkdir(exist_ok=True)
REPORT_DIR.mkdir(exist_ok=True)


def ensure_excel():
    if EXCEL_PATH.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily_Log"
    headers = [
        "日期", "数据来源", "行情分类", "T0参与度", "风险提示",
        "上证%", "创业板%", "三股平均涨幅%", "三股平均振幅%",
        "景旺最新价", "景旺涨跌幅%", "景旺振幅%",
        "中际最新价", "中际涨跌幅%", "中际振幅%",
        "胜宏最新价", "胜宏涨跌幅%", "胜宏振幅%",
        "非凸建议金额", "卡方建议金额", "皓兴建议金额", "跃然建议金额",
        "非凸实际收益", "卡方实际收益", "皓兴实际收益", "跃然实际收益", "备注"
    ]
    ws.append(headers)
    wb.save(EXCEL_PATH)


def append_to_excel(data, rec):
    ensure_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Daily_Log"]
    cls = rec["classification"]
    stocks = data.get("stocks", {})
    row = [
        data.get("date"), data.get("source"), cls["regime"], cls["t0_participation"], cls["risk"],
        data.get("market", {}).get("上证%"), data.get("market", {}).get("创业板%"),
        cls["avg_stock_change"], cls["avg_stock_amplitude"],
        stocks.get("景旺电子", {}).get("最新价"), stocks.get("景旺电子", {}).get("涨跌幅%"), stocks.get("景旺电子", {}).get("振幅%"),
        stocks.get("中际旭创", {}).get("最新价"), stocks.get("中际旭创", {}).get("涨跌幅%"), stocks.get("中际旭创", {}).get("振幅%"),
        stocks.get("胜宏科技", {}).get("最新价"), stocks.get("胜宏科技", {}).get("涨跌幅%"), stocks.get("胜宏科技", {}).get("振幅%"),
        rec["provider_amounts"].get("非凸"), rec["provider_amounts"].get("卡方"), rec["provider_amounts"].get("皓兴"), rec["provider_amounts"].get("跃然"),
        "", "", "", "", ""
    ]
    ws.append(row)
    wb.save(EXCEL_PATH)


def make_report(data, rec):
    cls = rec["classification"]
    html = f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><title>T0今日建议</title>
<style>body{{font-family:-apple-system,BlinkMacSystemFont,'PingFang SC',Arial;margin:40px;line-height:1.7}}table{{border-collapse:collapse}}td,th{{border:1px solid #ddd;padding:8px 12px}}th{{background:#f3f3f3}}</style></head>
<body>
<h1>T0今日建议 - {data.get('date')}</h1>
<h2>结论</h2>
<p><b>行情分类：</b>{cls['regime']}　<b>T0参与度：</b>{int(cls['t0_participation']*100)}%　<b>风险：</b>{cls['risk']}</p>
<p><b>建议参与资金：</b>{rec['active_amount']:,.0f} 元 / 持仓市值 {rec['total_position']:,.0f} 元</p>
<h2>四家量化分配</h2>
<table><tr><th>量化</th><th>建议金额</th><th>权重</th></tr>
"""
    for k, amount in rec["provider_amounts"].items():
        html += f"<tr><td>{k}</td><td>{amount:,.0f}</td><td>{rec['provider_weights'][k]*100:.0f}%</td></tr>"
    html += "</table><h2>按股票参与金额</h2><table><tr><th>股票</th><th>建议T0参与市值</th></tr>"
    for k, amount in rec["stock_amounts"].items():
        html += f"<tr><td>{k}</td><td>{amount:,.0f}</td></tr>"
    html += "</table><h2>个股行情</h2><table><tr><th>股票</th><th>最新价</th><th>涨跌幅%</th><th>振幅%</th><th>换手率%</th></tr>"
    for k, v in data.get("stocks", {}).items():
        html += f"<tr><td>{k}</td><td>{v.get('最新价')}</td><td>{v.get('涨跌幅%')}</td><td>{v.get('振幅%')}</td><td>{v.get('换手率%')}</td></tr>"
    html += "</table><p>收盘后请在 Excel 最右侧填入四家实际收益和备注。</p></body></html>"
    out = REPORT_DIR / f"report_{data.get('date')}.html"
    out.write_text(html, encoding="utf-8")
    return out


def main():
    print("正在抓取今日市场数据……")
    data = fetch_with_akshare(STOCKS)
    rec = recommend(data)
    append_to_excel(data, rec)
    report = make_report(data, rec)
    (DATA_DIR / f"market_{data.get('date')}.json").write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    cls = rec["classification"]
    print("\n===== T0 今日建议 =====")
    print(f"日期：{data.get('date')}")
    print(f"行情分类：{cls['regime']} | T0参与度：{int(cls['t0_participation']*100)}% | 风险：{cls['risk']}")
    print(f"建议参与资金：{rec['active_amount']:,.0f} 元")
    print("\n四家量化分配：")
    for k, amount in rec["provider_amounts"].items():
        print(f"- {k}: {amount:,.0f} 元")
    print(f"\nExcel已更新：{EXCEL_PATH}")
    print(f"报告已生成：{report}")


if __name__ == "__main__":
    main()
